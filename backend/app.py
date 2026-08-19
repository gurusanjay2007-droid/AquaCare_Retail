"""
PureFlow Service Hub – Flask Application
REST API Backend serving the SPA frontend
"""

import os
import json
from datetime import datetime, date, timedelta
from functools import wraps

import io
from flask import Flask, request, jsonify, send_from_directory, send_file, session

from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from flask_session import Session
from dotenv import load_dotenv

load_dotenv()

# ─── App Setup ───────────────────────────────────────────────
BASE_DIR  = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TMPL_DIR  = os.path.join(BASE_DIR, 'frontend', 'templates')
STATIC_DIR = os.path.join(BASE_DIR, 'frontend', 'static')

is_serverless = bool(os.getenv('VERCEL') or os.getenv('AWS_LAMBDA_FUNCTION_NAME'))

from werkzeug.middleware.proxy_fix import ProxyFix

app = Flask(__name__, template_folder=TMPL_DIR, static_folder=STATIC_DIR)

# WSGI Middleware to restore original request PATH_INFO from Vercel rewrite headers
class VercelPathMiddleware:
    def __init__(self, wsgi_app):
        self.wsgi_app = wsgi_app

    def __call__(self, environ, start_response):
        matched_path = environ.get('HTTP_X_MATCHED_PATH') or environ.get('HTTP_X_FORWARDED_URI') or environ.get('RAW_URI')
        if matched_path:
            clean_path = matched_path.split('?')[0]
            environ['PATH_INFO'] = clean_path
        return self.wsgi_app(environ, start_response)

app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_prefix=1)
app.wsgi_app = VercelPathMiddleware(app.wsgi_app)

app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'pureflow-dev-secret-2024')

# Database URI configuration with Supabase / PostgreSQL support
if is_serverless:
    sqlite_fallback = 'sqlite:////tmp/pureflow.db'
else:
    db_path = os.path.join(BASE_DIR, "pureflow.db")
    sqlite_fallback = f'sqlite:///{db_path}'

raw_db_url = os.getenv('DATABASE_URL', '').strip()
if not raw_db_url or '[YOUR-PASSWORD]' in raw_db_url or 'YOUR_PASSWORD' in raw_db_url:
    raw_db_url = sqlite_fallback
elif raw_db_url.startswith('postgres://'):
    raw_db_url = raw_db_url.replace('postgres://', 'postgresql://', 1)
elif is_serverless and raw_db_url.startswith('sqlite:'):
    raw_db_url = sqlite_fallback

app.config['SQLALCHEMY_DATABASE_URI'] = raw_db_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    'pool_pre_ping': True,
    'pool_recycle': 300,
}

# Session configuration:
# On serverless (Vercel), Flask's native signed client-side cookies provide stateless,
# resilient auth across ephemeral serverless instances and cold starts.
app.config['SESSION_COOKIE_NAME'] = 'aquacare_session'
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=7)

session_type = os.getenv('SESSION_TYPE')
if session_type:
    app.config['SESSION_TYPE'] = session_type
    if session_type == 'filesystem':
        app.config['SESSION_FILE_DIR'] = '/tmp/flask_session' if is_serverless else os.path.join(BASE_DIR, 'flask_session')
        try:
            os.makedirs(app.config['SESSION_FILE_DIR'], exist_ok=True)
        except Exception:
            pass
    try:
        Session(app)
    except Exception as e:
        print(f"Session init notice: {e}")

from backend.database import (
    db, User, Technician, Customer, Product,
    Installation, ServiceSchedule, Service,
    ServicePart, Bill, SmsLog, Inventory
)

db.init_app(app)


login_manager = LoginManager(app)
login_manager.login_view = 'serve_index'

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))


# ─── Helpers ─────────────────────────────────────────────────
def api_success(data=None, message='OK', status=200):
    return jsonify({'success': True, 'message': message, 'data': data}), status

def api_error(message='Error', status=400):
    return jsonify({'success': False, 'message': message}), status

def generate_invoice_number():
    today = date.today()
    count = Bill.query.filter(Bill.bill_date == today).count() + 1
    return f"INV-{today.strftime('%Y%m%d')}-{count:04d}"

def send_and_log_sms(customer, message):
    import urllib.request
    import urllib.parse
    import json
    
    api_key = os.getenv("SMS_API_KEY")
    status = 'Sent (Simulated)'
    
    # Auto-prepend country code for international SMS APIs (e.g. +91 for India)
    phone = customer.mobile.strip()
    if len(phone) == 10 and phone.isdigit():
        phone = "+91" + phone
        
    if api_key:
        params = {
            'authorization': api_key,
            'route': 'v3',
            'sender_id': 'TXTIND',
            'message': message,
            'language': 'english',
            'numbers': phone
        }
        url = "https://www.fast2sms.com/dev/bulkV2?" + urllib.parse.urlencode(params)
        try:
            req = urllib.request.Request(url, headers={'cache-control': 'no-cache'})
            with urllib.request.urlopen(req, timeout=8) as response:
                res_data = json.loads(response.read().decode('utf-8'))
                if res_data.get('return'):
                    status = 'Delivered'
                else:
                    status = f"Failed: {res_data.get('message', 'Unknown Error')}"
        except Exception as e:
            status = f"Error: {str(e)}"
    else:
        # Fallback to Textbelt free tier (1 free SMS per day per IP) for instant testing
        try:
            url = "https://textbelt.com/text"
            data_payload = urllib.parse.urlencode({
                'phone': phone,
                'message': message,
                'key': 'textbelt'
            }).encode('utf-8')
            req = urllib.request.Request(url, data=data_payload, method='POST')
            with urllib.request.urlopen(req, timeout=6) as response:
                res_data = json.loads(response.read().decode('utf-8'))
                if res_data.get('success'):
                    status = 'Delivered (Free Textbelt)'
                else:
                    status = f"Simulated (Textbelt limit: {res_data.get('error')})"
        except Exception as e:
            status = f"Simulated (Error: {str(e)})"
            
    sms = SmsLog(
        customer_id=customer.id,
        mobile=customer.mobile,
        message=message,
        status=status,
        sent_at=datetime.utcnow()
    )
    db.session.add(sms)
    return status in ['Delivered', 'Delivered (Free Textbelt)', 'Sent (Simulated)']

def log_sms(customer, message):
    return send_and_log_sms(customer, message)


# ─── Static / SPA ────────────────────────────────────────────
@app.route('/')
@app.route('/api')
@app.route('/api/index')
@app.route('/api/index.py')
def serve_index():
    return send_from_directory(TMPL_DIR, 'index.html')

@app.route('/favicon.ico')
def favicon():
    return send_from_directory(STATIC_DIR, 'logo.jpg', mimetype='image/jpeg')

@app.route('/logo.jpg')
def root_logo():
    return send_from_directory(STATIC_DIR, 'logo.jpg', mimetype='image/jpeg')

@app.route('/static/<path:filename>')
def serve_static(filename):
    return send_from_directory(STATIC_DIR, filename)


UPLOAD_FOLDER = '/tmp/uploads/installations' if is_serverless else os.path.join(STATIC_DIR, 'uploads', 'installations')
try:
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
except Exception:
    UPLOAD_FOLDER = '/tmp/uploads/installations'
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)

@app.route('/static/uploads/installations/<path:filename>')
def serve_installation_upload(filename):
    if os.path.exists(os.path.join(UPLOAD_FOLDER, filename)):
        return send_from_directory(UPLOAD_FOLDER, filename)
    static_fallback = os.path.join(STATIC_DIR, 'uploads', 'installations')
    if os.path.exists(os.path.join(static_fallback, filename)):
        return send_from_directory(static_fallback, filename)
    return ('File not found', 404)

@app.errorhandler(404)
def handle_404(e):
    if request.path.startswith('/api/') and not request.path.startswith('/api/index'):
        return jsonify({'success': False, 'message': 'API endpoint not found'}), 404
    return send_from_directory(TMPL_DIR, 'index.html')


# ═══════════════════════════════════════════════════════════════
# FILE UPLOAD ROUTE
# ═══════════════════════════════════════════════════════════════
@app.route('/api/upload', methods=['POST'])
def upload_file():
    if 'photo' not in request.files:
        return api_error('No photo file provided')
    file = request.files['photo']
    if file.filename == '':
        return api_error('No selected file')
    if file:
        ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else 'jpg'
        filename = f"inst_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{os.urandom(4).hex()}.{ext}"
        filepath = os.path.join(UPLOAD_FOLDER, filename)
        file.save(filepath)
        photo_url = f"/static/uploads/installations/{filename}"
        return api_success({'photo_url': photo_url}, 'File uploaded successfully')
    return api_error('Failed to upload file')


# ═══════════════════════════════════════════════════════════════
# AUTH ROUTES
# ═══════════════════════════════════════════════════════════════
@app.route('/api/auth/register', methods=['POST'])
def register():
    data = request.get_json()
    required = ['business_name','owner_name','email','mobile','password','business_address']
    for f in required:
        if not data.get(f):
            return api_error(f'Field "{f}" is required')

    if User.query.filter_by(email=data['email']).first():
        return api_error('Email already registered', 409)

    user = User(
        business_name    = data['business_name'],
        owner_name       = data['owner_name'],
        email            = data['email'],
        mobile           = data['mobile'],
        business_address = data['business_address'],
        gst_number       = data.get('gst_number')
    )
    user.set_password(data['password'])
    db.session.add(user)
    db.session.commit()
    session['user_role'] = 'admin'
    login_user(user)
    res_data = user.to_dict()
    res_data['role'] = 'admin'
    return api_success(res_data, 'Registered successfully', 201)


@app.route('/api/auth/login', methods=['POST'])
def login():
    data = request.get_json() or {}
    email = data.get('email', '').strip()
    password = data.get('password', '')
    user = User.query.filter_by(email=email).first()
    if not user or not user.check_password(password):
        return api_error('Invalid email or password', 401)
    session['user_role'] = 'admin'
    login_user(user, remember=True)
    res_data = user.to_dict()
    res_data['role'] = 'admin'
    return api_success(res_data, 'Logged in')


@app.route('/api/auth/tech-login', methods=['POST'])
def tech_login():
    data = request.get_json() or {}
    tech_id = data.get('tech_id', '').strip()
    passcode = data.get('passcode', '').strip()
    if not tech_id or not passcode:
        return api_error('Tech ID and passcode are required')

    t = Technician.query.filter_by(tech_id=tech_id).first()
    if not t or not t.check_passcode(passcode):
        return api_error('Invalid Technician ID or Passcode', 401)
    if t.status != 'Active':
        return api_error('Technician account is inactive. Please contact Hub Manager.', 403)

    session['tech_db_id'] = t.id
    session['tech_id'] = t.tech_id
    session['user_role'] = 'technician'

    res_data = t.to_dict()
    res_data['role'] = 'technician'
    return api_success(res_data, 'Logged in as Technician')


@app.route('/api/auth/logout', methods=['POST'])
def logout():
    logout_user()
    session.clear()
    return api_success(message='Logged out')


@app.route('/api/auth/me', methods=['GET'])
def me():
    if current_user.is_authenticated:
        res_data = current_user.to_dict()
        res_data['role'] = 'admin'
        return api_success(res_data)
    elif session.get('user_role') == 'technician' and session.get('tech_db_id'):
        t = Technician.query.get(session.get('tech_db_id'))
        if t:
            res_data = t.to_dict()
            res_data['role'] = 'technician'
            return api_success(res_data)
    return api_error('Not authenticated', 401)


# ═══════════════════════════════════════════════════════════════
# DASHBOARD
# ═══════════════════════════════════════════════════════════════
@app.route('/api/dashboard', methods=['GET'])
def dashboard():
    today = date.today()
    thirty_days_ago = today - timedelta(days=30)

    total_customers     = Customer.query.count()
    total_installations = Installation.query.count()
    total_services      = Service.query.count()
    total_technicians   = Technician.query.filter_by(status='Active').count()

    # Revenue & profit
    all_bills     = Bill.query.all()
    total_revenue = sum(float(b.grand_total) for b in all_bills)
    paid_revenue  = sum(float(b.grand_total) for b in all_bills if b.payment_status == 'Paid')

    # Installation profit
    all_inst      = Installation.query.all()
    install_profit= sum(i.profit for i in all_inst)

    # Service profit from parts
    all_parts     = ServicePart.query.all()
    parts_profit  = sum(p.profit for p in all_parts)
    total_profit  = install_profit + parts_profit

    # Overdue services
    overdue_count = ServiceSchedule.query.filter(
        ServiceSchedule.status != 'Completed',
        ServiceSchedule.next_service_date < today
    ).count()

    # Update overdue statuses
    overdue_schedules = ServiceSchedule.query.filter(
        ServiceSchedule.status == 'Pending',
        ServiceSchedule.next_service_date < today
    ).all()
    for s in overdue_schedules:
        s.status = 'Overdue'
    if overdue_schedules:
        db.session.commit()

    # Pending services (due within 7 days)
    due_soon = ServiceSchedule.query.filter(
        ServiceSchedule.status == 'Pending',
        ServiceSchedule.next_service_date <= today + timedelta(days=7),
        ServiceSchedule.next_service_date >= today
    ).count()

    # Low stock items
    low_stock_count = sum(1 for i in Inventory.query.all() if i.is_low_stock)

    # ── Daily Service Box Summary & List for Today ──
    today_completed_services = Service.query.filter_by(service_date=today).all()
    today_pending_schedules  = ServiceSchedule.query.filter_by(next_service_date=today, status='Pending').all()
    all_overdue_schedules    = ServiceSchedule.query.filter(
        ServiceSchedule.status == 'Overdue',
        ServiceSchedule.next_service_date < today
    ).all()

    today_completed_count = len(today_completed_services)
    today_pending_count   = len(today_pending_schedules)
    today_overdue_count   = len(all_overdue_schedules)
    today_total_count     = today_completed_count + today_pending_count + today_overdue_count

    # Daily service list items
    daily_service_list = []
    for s in today_completed_services:
        daily_service_list.append({
            'type': 'Service',
            'id': s.id,
            'customer_name': s.customer.customer_name if s.customer else 'N/A',
            'customer_mobile': s.customer.mobile if s.customer else 'N/A',
            'technician_name': s.technician.technician_name if s.technician else 'N/A',
            'service_type': s.service_type,
            'status': 'Completed',
            'amount': float(s.total_bill or 0),
            'date': s.service_date.isoformat()
        })
    for sc in today_pending_schedules:
        daily_service_list.append({
            'type': 'Schedule',
            'id': sc.id,
            'customer_name': sc.customer.customer_name if sc.customer else 'N/A',
            'customer_mobile': sc.customer.mobile if sc.customer else 'N/A',
            'technician_name': 'Pending Assignment',
            'service_type': 'Scheduled Visit',
            'status': 'Pending',
            'amount': 0.0,
            'date': sc.next_service_date.isoformat()
        })
    for sc in all_overdue_schedules:
        daily_service_list.append({
            'type': 'Schedule',
            'id': sc.id,
            'customer_name': sc.customer.customer_name if sc.customer else 'N/A',
            'customer_mobile': sc.customer.mobile if sc.customer else 'N/A',
            'technician_name': 'Pending Assignment',
            'service_type': 'Overdue Service',
            'status': 'Overdue',
            'amount': 0.0,
            'date': sc.next_service_date.isoformat()
        })

    # Recent services (last 30 days)
    recent_services = Service.query.filter(
        Service.service_date >= thirty_days_ago
    ).order_by(Service.service_date.desc()).limit(5).all()

    # Monthly revenue chart data (last 6 months)
    monthly_data = []
    for i in range(5, -1, -1):
        month_start = (today.replace(day=1) - timedelta(days=30*i))
        month_start = month_start.replace(day=1)
        if month_start.month == 12:
            month_end = month_start.replace(year=month_start.year+1, month=1, day=1)
        else:
            month_end = month_start.replace(month=month_start.month+1, day=1)
        month_bills = Bill.query.filter(
            Bill.bill_date >= month_start,
            Bill.bill_date < month_end
        ).all()
        monthly_data.append({
            'month': month_start.strftime('%b %Y'),
            'revenue': sum(float(b.grand_total) for b in month_bills),
            'paid': sum(float(b.grand_total) for b in month_bills if b.payment_status == 'Paid'),
        })

    # Service type breakdown
    service_types = {}
    for s in Service.query.all():
        service_types[s.service_type] = service_types.get(s.service_type, 0) + 1

    return api_success({
        'stats': {
            'total_customers': total_customers,
            'total_installations': total_installations,
            'total_services': total_services,
            'total_technicians': total_technicians,
            'total_revenue': round(total_revenue, 2),
            'paid_revenue': round(paid_revenue, 2),
            'total_profit': round(total_profit, 2),
            'overdue_count': overdue_count,
            'due_soon': due_soon,
            'low_stock_count': low_stock_count,
            'daily_services': {
                'total': today_total_count,
                'completed': today_completed_count,
                'pending': today_pending_count,
                'overdue': today_overdue_count,
                'list': daily_service_list
            }
        },
        'monthly_revenue': monthly_data,
        'service_types': service_types,
        'recent_services': [s.to_dict() for s in recent_services],
    })



# ═══════════════════════════════════════════════════════════════
# CUSTOMERS
# ═══════════════════════════════════════════════════════════════
@app.route('/api/customers', methods=['GET'])
def get_customers():
    q      = request.args.get('q', '')
    page   = int(request.args.get('page', 1))
    limit  = int(request.args.get('limit', 20))
    query  = Customer.query
    if q:
        query = query.filter(
            db.or_(Customer.customer_name.ilike(f'%{q}%'),
                   Customer.mobile.ilike(f'%{q}%'),
                   Customer.city.ilike(f'%{q}%'))
        )
    total  = query.count()
    items  = query.order_by(Customer.created_at.desc()).offset((page-1)*limit).limit(limit).all()
    return api_success({'items': [c.to_dict() for c in items], 'total': total})


@app.route('/api/customers/<int:cid>', methods=['GET'])
def get_customer(cid):
    c = Customer.query.get_or_404(cid)
    data = c.to_dict()
    data['installations'] = [i.to_dict() for i in c.installations]
    data['services']      = [s.to_dict() for s in c.services]
    data['bills']         = [b.to_dict() for b in c.bills]
    data['sms_logs']      = [l.to_dict() for l in c.sms_logs]
    data['schedule']      = [sc.to_dict() for sc in c.service_schedule]
    return api_success(data)


@app.route('/api/customers', methods=['POST'])
def create_customer():
    data = request.get_json()
    if not data.get('customer_name') or not data.get('mobile'):
        return api_error('customer_name and mobile are required')
    if Customer.query.filter_by(mobile=data['mobile']).first():
        return api_error('Mobile number already registered', 409)
    c = Customer(**{k: data.get(k) for k in [
        'customer_name','mobile','alternate_mobile','address','landmark','city','pincode'
    ]})
    if current_user.is_authenticated:
        c.user_id = current_user.id
    db.session.add(c)
    db.session.commit()
    return api_success(c.to_dict(), 'Customer created', 201)


@app.route('/api/customers/<int:cid>', methods=['PUT'])
def update_customer(cid):
    c    = Customer.query.get_or_404(cid)
    data = request.get_json()
    for field in ['customer_name','mobile','alternate_mobile','address','landmark','city','pincode']:
        if field in data:
            setattr(c, field, data[field])
    db.session.commit()
    return api_success(c.to_dict(), 'Customer updated')


@app.route('/api/customers/<int:cid>', methods=['DELETE'])
def delete_customer(cid):
    c = Customer.query.get_or_404(cid)
    db.session.delete(c)
    db.session.commit()
    return api_success(message='Customer deleted')


# ═══════════════════════════════════════════════════════════════
# TECHNICIANS
# ═══════════════════════════════════════════════════════════════
@app.route('/api/technicians', methods=['GET'])
def get_technicians():
    items = Technician.query.order_by(Technician.technician_name).all()
    return api_success([t.to_dict() for t in items])


@app.route('/api/technicians', methods=['POST'])
def create_technician():
    data = request.get_json()
    if not data.get('technician_name') or not data.get('mobile'):
        return api_error('technician_name and mobile are required')
    
    tech_id = data.get('tech_id')
    if not tech_id:
        count = Technician.query.count() + 101
        tech_id = f"TECH{count}"

    if Technician.query.filter_by(tech_id=tech_id).first():
        return api_error(f'Technician ID "{tech_id}" already exists', 409)

    t = Technician(
        technician_name = data['technician_name'],
        mobile          = data['mobile'],
        email           = data.get('email'),
        address         = data.get('address'),
        status          = data.get('status', 'Active'),
        tech_id         = tech_id,
        passcode        = data.get('passcode', '123456')
    )
    db.session.add(t)
    db.session.commit()
    return api_success(t.to_dict(), 'Technician created', 201)


@app.route('/api/technicians/<int:tid>', methods=['PUT'])
def update_technician(tid):
    t    = Technician.query.get_or_404(tid)
    data = request.get_json()
    for field in ['technician_name','mobile','email','address','status','tech_id','passcode','photo_url']:
        if field in data and data[field] is not None:
            setattr(t, field, data[field])
    db.session.commit()
    return api_success(t.to_dict(), 'Technician updated')


@app.route('/api/technicians/<int:tid>', methods=['DELETE'])
def delete_technician(tid):
    t = Technician.query.get_or_404(tid)
    db.session.delete(t)
    db.session.commit()
    return api_success(message='Technician deleted')


# ═══════════════════════════════════════════════════════════════
# PRODUCTS
# ═══════════════════════════════════════════════════════════════
@app.route('/api/products', methods=['GET'])
def get_products():
    items = Product.query.order_by(Product.product_name).all()
    return api_success([p.to_dict() for p in items])


@app.route('/api/products', methods=['POST'])
def create_product():
    data = request.get_json()
    if not data.get('product_name') or not data.get('brand'):
        return api_error('product_name and brand are required')
    p = Product(
        product_name    = data.get('product_name'),
        brand           = data.get('brand'),
        model_number    = data.get('model_number'),
        serial_number   = data.get('serial_number'),
        cost_price      = float(data.get('cost_price', 0.0)),
        selling_price   = float(data.get('selling_price', 0.0)),
        warranty_months = int(data.get('warranty_months', 12))
    )
    db.session.add(p)
    db.session.commit()
    return api_success(p.to_dict(), 'Product created', 201)


@app.route('/api/products/<int:pid>', methods=['PUT'])
def update_product(pid):
    p    = Product.query.get_or_404(pid)
    data = request.get_json()
    for field in ['product_name','brand','model_number','serial_number','warranty_months','cost_price','selling_price']:
        if field in data:
            if field in ['cost_price', 'selling_price']:
                setattr(p, field, float(data[field]))
            elif field == 'warranty_months':
                setattr(p, field, int(data[field]))
            else:
                setattr(p, field, data[field])
    db.session.commit()
    return api_success(p.to_dict(), 'Product updated')



# ═══════════════════════════════════════════════════════════════
# INSTALLATIONS
# ═══════════════════════════════════════════════════════════════
@app.route('/api/installations', methods=['GET'])
def get_installations():
    q     = request.args.get('q', '')
    page  = int(request.args.get('page', 1))
    limit = int(request.args.get('limit', 20))
    query = Installation.query.join(Customer)
    if q:
        query = query.filter(
            db.or_(Customer.customer_name.ilike(f'%{q}%'),
                   Customer.mobile.ilike(f'%{q}%'))
        )
    total = query.count()
    items = query.order_by(Installation.installation_date.desc()).offset((page-1)*limit).limit(limit).all()
    return api_success({'items': [i.to_dict() for i in items], 'total': total})


@app.route('/api/installations', methods=['POST'])
def create_installation():
    data = request.get_json() or {}

    # Handle customer creation on-the-fly if customer_id not passed directly
    customer_id = data.get('customer_id')
    if not customer_id:
        if data.get('customer_name') and data.get('mobile'):
            mobile = data['mobile'].strip()
            existing = Customer.query.filter_by(mobile=mobile).first()
            if existing:
                customer_id = existing.id
            else:
                c = Customer(
                    customer_name = data['customer_name'].strip(),
                    mobile        = mobile,
                    address       = data.get('address'),
                    city          = data.get('city', 'Hyderabad')
                )
                if current_user.is_authenticated:
                    c.user_id = current_user.id
                db.session.add(c)
                db.session.flush()
                customer_id = c.id
        else:
            return api_error('Customer ID or Customer Name & Mobile is required')

    technician_id = data.get('technician_id')
    if not technician_id and session.get('tech_db_id'):
        technician_id = session.get('tech_db_id')

    if not customer_id or not data.get('product_id') or not technician_id:
        return api_error('customer, product, and technician are required')

    product = Product.query.get(data['product_id'])
    cost_price = float(data.get('cost_price')) if data.get('cost_price') is not None else (float(product.cost_price) if product else 0.0)
    selling_price = float(data.get('selling_price')) if data.get('selling_price') is not None else (float(product.selling_price) if product else 0.0)

    inst_date_str = data.get('installation_date')
    if inst_date_str:
        inst_date = datetime.strptime(inst_date_str, '%Y-%m-%d').date()
    else:
        inst_date = date.today()

    inst = Installation(
        customer_id       = customer_id,
        product_id        = data['product_id'],
        technician_id     = technician_id,
        installation_photo= data.get('installation_photo'),
        source_water_type = data.get('source_water_type', 'Municipal'),
        input_tds         = data.get('input_tds'),
        output_tds        = data.get('output_tds'),
        installation_date = inst_date,
        cost_price        = cost_price,
        selling_price     = selling_price,
        remarks           = data.get('remarks'),
    )
    db.session.add(inst)
    db.session.flush()

    # Auto-create service schedule (6 months default or user warranty/interval)
    interval = int(data.get('service_interval_months', 6))
    next_date = inst.installation_date + timedelta(days=30 * interval)
    schedule  = ServiceSchedule(
        installation_id         = inst.id,
        customer_id             = inst.customer_id,
        next_service_date       = next_date,
        service_interval_months = interval,
        status                  = 'Pending',
        reminder_sent           = 'No',
    )
    db.session.add(schedule)
    db.session.commit()
    return api_success(inst.to_dict(), 'Installation recorded with service schedule', 201)


@app.route('/api/installations/<int:iid>', methods=['GET'])
def get_installation(iid):
    inst = Installation.query.get_or_404(iid)
    data = inst.to_dict()
    data['services']  = [s.to_dict() for s in inst.services]
    data['schedule']  = [sc.to_dict() for sc in inst.service_schedule]
    return api_success(data)


@app.route('/api/installations/<int:iid>', methods=['PUT'])
def update_installation(iid):
    inst = Installation.query.get_or_404(iid)
    data = request.get_json()
    for field in ['source_water_type','input_tds','output_tds','cost_price','selling_price','remarks','installation_photo']:
        if field in data:
            setattr(inst, field, data[field])
    if 'installation_date' in data:
        inst.installation_date = datetime.strptime(data['installation_date'], '%Y-%m-%d').date()
    db.session.commit()
    return api_success(inst.to_dict(), 'Installation updated')


# ═══════════════════════════════════════════════════════════════
# SERVICE SCHEDULES
# ═══════════════════════════════════════════════════════════════
@app.route('/api/schedules', methods=['GET'])
def get_schedules():
    status = request.args.get('status', '')
    q      = request.args.get('q', '').strip()
    today  = date.today()

    # Auto-update overdue
    overdue = ServiceSchedule.query.filter(
        ServiceSchedule.status == 'Pending',
        ServiceSchedule.next_service_date < today
    ).all()
    for s in overdue:
        s.status = 'Overdue'
    if overdue:
        db.session.commit()

    query = ServiceSchedule.query.join(Customer)
    if status:
        query = query.filter(ServiceSchedule.status == status)
    if q:
        query = query.filter(
            db.or_(Customer.customer_name.ilike(f'%{q}%'),
                   Customer.mobile.ilike(f'%{q}%'))
        )
    items = query.order_by(ServiceSchedule.next_service_date).all()
    return api_success([s.to_dict() for s in items])


@app.route('/api/schedules/<int:sid>', methods=['PUT'])
def update_schedule(sid):
    s    = ServiceSchedule.query.get_or_404(sid)
    data = request.get_json()
    
    # Trigger actual SMS if manual reminder sent via UI
    if data.get('reminder_sent') == 'Yes' and s.reminder_sent != 'Yes':
        customer = s.customer
        if customer:
            msg = (
                f"Dear {customer.customer_name}, your Aqua Care RO Service is scheduled for "
                f"{s.next_service_date.strftime('%d-%b-%Y')}. Please confirm your availability. Thank you!"
            )
            send_and_log_sms(customer, msg)
            
    is_marking_completed = data.get('status') == 'Completed' and s.status != 'Completed'

    for field in ['status','reminder_sent','next_service_date','service_interval_months']:
        if field in data:
            if field == 'next_service_date':
                s.next_service_date = datetime.strptime(data[field], '%Y-%m-%d').date()
            else:
                setattr(s, field, data[field])

    next_date_str = None
    if is_marking_completed:
        # Auto-reappoint next service schedule
        interval = s.service_interval_months or 6
        next_date = date.today() + timedelta(days=30 * interval)
        new_sched = ServiceSchedule(
            installation_id         = s.installation_id,
            customer_id             = s.customer_id,
            next_service_date       = next_date,
            service_interval_months = interval,
            status                  = 'Pending',
            reminder_sent           = 'No',
        )
        db.session.add(new_sched)
        next_date_str = next_date.strftime('%d-%b-%Y')

        # Create actual Service record for today
        tech = Technician.query.filter_by(status='Active').first()
        if not tech:
            tech = Technician.query.first()
        tech_id = tech.id if tech else 1

        svc = Service(
            installation_id = s.installation_id,
            customer_id     = s.customer_id,
            technician_id   = tech_id,
            service_date    = date.today(),
            service_type    = 'Regular',
            service_charge  = 0.00,
            remarks         = 'Marked completed from schedules checklist'
        )
        db.session.add(svc)

    db.session.commit()

    res_data = s.to_dict()
    if next_date_str:
        msg = f"Service marked completed! Reappointed for next visit on {next_date_str}."
    else:
        msg = "Schedule updated."

    return api_success(res_data, msg)



# ═══════════════════════════════════════════════════════════════
# SERVICES
# ═══════════════════════════════════════════════════════════════
@app.route('/api/services', methods=['GET'])
def get_services():
    q     = request.args.get('q', '')
    page  = int(request.args.get('page', 1))
    limit = int(request.args.get('limit', 20))
    query = Service.query.join(Customer)
    if q:
        query = query.filter(
            db.or_(Customer.customer_name.ilike(f'%{q}%'),
                   Customer.mobile.ilike(f'%{q}%'))
        )
    total = query.count()
    items = query.order_by(Service.service_date.desc()).offset((page-1)*limit).limit(limit).all()
    return api_success({'items': [s.to_dict() for s in items], 'total': total})


@app.route('/api/services', methods=['POST'])
def create_service():
    data = request.get_json() or {}

    technician_id = data.get('technician_id')
    if not technician_id and session.get('tech_db_id'):
        technician_id = session.get('tech_db_id')

    customer_id = data.get('customer_id')
    installation_id = data.get('installation_id')

    if not installation_id and customer_id:
        inst = Installation.query.filter_by(customer_id=customer_id).order_by(Installation.installation_date.desc()).first()
        if inst:
            installation_id = inst.id

    if not customer_id and installation_id:
        inst = Installation.query.get(installation_id)
        if inst:
            customer_id = inst.customer_id

    if not installation_id or not customer_id or not technician_id:
        return api_error('customer, installation, and technician are required')

    svc_date_str = data.get('service_date')
    if svc_date_str:
        svc_date = datetime.strptime(svc_date_str, '%Y-%m-%d').date()
    else:
        svc_date = date.today()

    svc = Service(
        installation_id = installation_id,
        customer_id     = customer_id,
        technician_id   = technician_id,
        service_date    = svc_date,
        service_type    = data.get('service_type', 'Regular'),
        tds_before      = data.get('tds_before'),
        tds_after       = data.get('tds_after'),
        service_charge  = float(data.get('service_charge', 0)),
        remarks         = data.get('remarks'),
    )

    # Add parts if provided
    parts_data = data.get('parts', [])
    total_parts_cost = 0.0
    for pd in parts_data:
        qty       = int(pd.get('quantity', 1))
        sp_price  = float(pd.get('selling_price', 0))
        cp_price  = float(pd.get('cost_price', 0))
        part = ServicePart(
            part_name     = pd['part_name'],
            quantity      = qty,
            cost_price    = cp_price,
            selling_price = sp_price,
        )
        svc.parts.append(part)
        total_parts_cost += sp_price * qty

        # Optional inventory deduction if item exists
        inv = Inventory.query.filter(
            Inventory.part_name.ilike(pd['part_name'])
        ).first()
        if inv and inv.available_stock >= qty:
            inv.available_stock -= qty

    svc.parts_total_cost = total_parts_cost
    svc.total_bill       = float(svc.service_charge) + total_parts_cost

    db.session.add(svc)

    next_service_date_str = None
    # Mark schedule as completed
    schedule = ServiceSchedule.query.filter_by(
        installation_id=data['installation_id'],
        status='Pending'
    ).order_by(ServiceSchedule.next_service_date).first()
    if not schedule:
        schedule = ServiceSchedule.query.filter_by(
            installation_id=data['installation_id'],
            status='Overdue'
        ).order_by(ServiceSchedule.next_service_date).first()

    if schedule:
        schedule.status = 'Completed'
        # Create next schedule
        interval  = schedule.service_interval_months
        next_date = svc.service_date + timedelta(days=30 * interval)
        new_sched = ServiceSchedule(
            installation_id         = data['installation_id'],
            customer_id             = data['customer_id'],
            next_service_date       = next_date,
            service_interval_months = interval,
            status                  = 'Pending',
            reminder_sent           = 'No',
        )
        db.session.add(new_sched)
        next_service_date_str = next_date.isoformat()

    db.session.commit()

    res_dict = svc.to_dict()
    if next_service_date_str:
        res_dict['next_service_date'] = next_service_date_str
        msg = f"Service marked complete! Next Service scheduled for {next_date.strftime('%d-%b-%Y')}."
    else:
        msg = "Service recorded successfully."

    return api_success(res_dict, msg, 201)



@app.route('/api/services/<int:sid>', methods=['GET'])
def get_service(sid):
    s = Service.query.get_or_404(sid)
    return api_success(s.to_dict())


# ═══════════════════════════════════════════════════════════════
# BILLS
# ═══════════════════════════════════════════════════════════════
@app.route('/api/bills', methods=['GET'])
def get_bills():
    q     = request.args.get('q', '')
    page  = int(request.args.get('page', 1))
    limit = int(request.args.get('limit', 20))
    query = Bill.query.join(Customer)
    if q:
        query = query.filter(
            db.or_(Customer.customer_name.ilike(f'%{q}%'),
                   Bill.invoice_number.ilike(f'%{q}%'))
        )
    total = query.count()
    items = query.order_by(Bill.bill_date.desc()).offset((page-1)*limit).limit(limit).all()
    return api_success({'items': [b.to_dict() for b in items], 'total': total})


@app.route('/api/bills', methods=['POST'])
def create_bill():
    data = request.get_json()
    if not data.get('customer_id'):
        return api_error('customer_id is required')

    customer = Customer.query.get_or_404(data['customer_id'])
    bill = Bill(
        invoice_number  = generate_invoice_number(),
        customer_id     = data['customer_id'],
        installation_id = data.get('installation_id'),
        service_id      = data.get('service_id'),
        bill_date       = datetime.strptime(data.get('bill_date', date.today().isoformat()), '%Y-%m-%d').date(),
        subtotal        = float(data.get('subtotal', 0)),
        service_charge  = float(data.get('service_charge', 0)),
        grand_total     = float(data.get('grand_total', 0)),
        payment_status  = data.get('payment_status', 'Unpaid'),
        payment_mode    = data.get('payment_mode', 'Cash'),
    )
    db.session.add(bill)
    db.session.flush()

    # Auto log SMS
    sms_msg = (
        f"Dear {customer.customer_name}, your invoice {bill.invoice_number} "
        f"of Rs.{bill.grand_total:.2f} has been generated. "
        f"Payment Status: {bill.payment_status}. "
        f"Thank you – Aqua Care Water Solutions."
    )
    log_sms(customer, sms_msg)
    db.session.commit()
    return api_success(bill.to_dict(), 'Bill created', 201)


@app.route('/api/bills/<int:bid>', methods=['PUT'])
def update_bill(bid):
    b    = Bill.query.get_or_404(bid)
    data = request.get_json()
    for field in ['payment_status','payment_mode','subtotal','service_charge','grand_total']:
        if field in data:
            setattr(b, field, data[field])
    db.session.commit()
    return api_success(b.to_dict(), 'Bill updated')


# ═══════════════════════════════════════════════════════════════
# INVENTORY
# ═══════════════════════════════════════════════════════════════
@app.route('/api/inventory', methods=['GET'])
def get_inventory():
    low_only = request.args.get('low_stock') == 'true'
    q        = request.args.get('q', '').strip()
    items    = Inventory.query.order_by(Inventory.part_name).all()
    if low_only:
        items = [i for i in items if i.is_low_stock]
    if q:
        items = [i for i in items if q.lower() in (i.part_name or '').lower() or (i.brand and q.lower() in i.brand.lower())]
    return api_success([i.to_dict() for i in items])


@app.route('/api/inventory', methods=['POST'])
def create_inventory():
    data = request.get_json()
    if not data.get('part_name'):
        return api_error('part_name is required')
    inv = Inventory(**{k: data.get(k) for k in [
        'part_name','brand','available_stock','purchase_price','selling_price','reorder_level'
    ] if data.get(k) is not None})
    db.session.add(inv)
    db.session.commit()
    return api_success(inv.to_dict(), 'Inventory item created', 201)


@app.route('/api/inventory/<int:iid>', methods=['PUT'])
def update_inventory(iid):
    inv  = Inventory.query.get_or_404(iid)
    data = request.get_json()
    for field in ['part_name','brand','available_stock','purchase_price','selling_price','reorder_level']:
        if field in data:
            setattr(inv, field, data[field])
    db.session.commit()
    return api_success(inv.to_dict(), 'Inventory updated')


@app.route('/api/inventory/<int:iid>', methods=['DELETE'])
def delete_inventory(iid):
    inv = Inventory.query.get_or_404(iid)
    db.session.delete(inv)
    db.session.commit()
    return api_success(message='Item deleted')


# ═══════════════════════════════════════════════════════════════
# SMS LOGS
# ═══════════════════════════════════════════════════════════════
@app.route('/api/sms-logs', methods=['GET'])
def get_sms_logs():
    page  = int(request.args.get('page', 1))
    limit = int(request.args.get('limit', 20))
    q     = request.args.get('q', '').strip()
    query = SmsLog.query
    if q:
        query = query.filter(
            db.or_(SmsLog.mobile.ilike(f'%{q}%'),
                   SmsLog.message.ilike(f'%{q}%'),
                   SmsLog.status.ilike(f'%{q}%'))
        )
    total = query.count()
    items = query.order_by(SmsLog.sent_at.desc()).offset((page-1)*limit).limit(limit).all()
    return api_success({'items': [s.to_dict() for s in items], 'total': total})


# ═══════════════════════════════════════════════════════════════
# CRON / AUTOMATED REMINDERS
# ═══════════════════════════════════════════════════════════════
@app.route('/api/cron/send-reminders', methods=['GET', 'POST'])
def cron_send_reminders():
    cron_secret = os.getenv('CRON_SECRET')
    if cron_secret and request.headers.get('Authorization') != f"Bearer {cron_secret}":
        return api_error('Unauthorized cron request', 401)

    today = date.today()
    target_date = today + timedelta(days=3)  # Remind 3 days in advance
    
    upcoming_schedules = ServiceSchedule.query.filter(
        ServiceSchedule.status.in_(['Pending', 'Overdue']),
        ServiceSchedule.next_service_date <= target_date,
        ServiceSchedule.reminder_sent == 'No'
    ).all()

    sent_count = 0
    for s in upcoming_schedules:
        customer = s.customer
        if not customer:
            continue
        
        msg = (
            f"Dear {customer.customer_name}, your Aqua Care RO Service is due on "
            f"{s.next_service_date.strftime('%d-%b-%Y')}. Our technician will contact you shortly. Thank you!"
        )
        success = send_and_log_sms(customer, msg)
        if success:
            s.reminder_sent = 'Yes'
            sent_count += 1
            
    db.session.commit()
    return api_success({'reminders_sent': sent_count}, f"Automated reminder scan complete. Sent {sent_count} reminders.")


def parse_report_dates(filter_type, args):
    today = date.today()
    if filter_type == 'all':
        return date(2000, 1, 1), date(2099, 12, 31)
    elif filter_type == 'single_date':
        dt_str = args.get('date')
        d = datetime.strptime(dt_str, '%Y-%m-%d').date() if dt_str else today
        return d, d
    elif filter_type == 'week':
        dt_str = args.get('date')
        ref = datetime.strptime(dt_str, '%Y-%m-%d').date() if dt_str else today
        s = ref - timedelta(days=ref.weekday())
        e = s + timedelta(days=6)
        return s, e
    elif filter_type == 'month':
        m_str = args.get('month')
        if m_str:
            y, m = map(int, m_str.split('-'))
            s = date(y, m, 1)
            if m == 12:
                e = date(y + 1, 1, 1) - timedelta(days=1)
            else:
                e = date(y, m + 1, 1) - timedelta(days=1)
            return s, e
        return date(today.year, today.month, 1), today
    else: # custom
        s_str = args.get('start_date')
        e_str = args.get('end_date')
        s = datetime.strptime(s_str, '%Y-%m-%d').date() if s_str else date(2020, 1, 1)
        e = datetime.strptime(e_str, '%Y-%m-%d').date() if e_str else today
        return s, e


@app.route('/api/reports/preview', methods=['GET'])
def preview_report():
    report_type = request.args.get('report_type', 'installations')
    filter_type = request.args.get('filter_type', 'all')
    
    start_date, end_date = parse_report_dates(filter_type, request.args)
    
    if report_type == 'installations':
        items = Installation.query.filter(Installation.installation_date >= start_date, Installation.installation_date <= end_date).order_by(Installation.installation_date.desc()).all()
        rows = []
        for i in items:
            c = i.customer
            p = i.product
            t = i.technician
            rows.append({
                'id': i.id,
                'customer_name': c.customer_name if c else 'N/A',
                'mobile': c.mobile if c else 'N/A',
                'product_name': p.product_name if p else 'N/A',
                'technician_name': t.technician_name if t else 'N/A',
                'date': i.installation_date.strftime('%Y-%m-%d'),
                'input_tds': i.input_tds,
                'output_tds': i.output_tds,
                'selling_price': float(i.selling_price or 0)
            })
        return api_success({'type': 'installations', 'rows': rows, 'total': len(rows)})
    elif report_type == 'services':
        items = Service.query.filter(Service.service_date >= start_date, Service.service_date <= end_date).order_by(Service.service_date.desc()).all()
        rows = []
        for s in items:
            c = s.customer
            t = s.technician
            rows.append({
                'id': s.id,
                'customer_name': c.customer_name if c else 'N/A',
                'mobile': c.mobile if c else 'N/A',
                'service_type': s.service_type,
                'technician_name': t.technician_name if t else 'N/A',
                'date': s.service_date.strftime('%Y-%m-%d'),
                'tds_before': s.tds_before,
                'tds_after': s.tds_after,
                'total_bill': float(s.total_bill or 0)
            })
        return api_success({'type': 'services', 'rows': rows, 'total': len(rows)})
    else: # master
        inst_count = Installation.query.filter(Installation.installation_date >= start_date, Installation.installation_date <= end_date).count()
        serv_count = Service.query.filter(Service.service_date >= start_date, Service.service_date <= end_date).count()
        rev = sum(float(b.grand_total) for b in Bill.query.filter(Bill.bill_date >= start_date, Bill.bill_date <= end_date).all())
        rows = [
            {'category': 'Summary', 'metric': 'Total Installations', 'val': inst_count},
            {'category': 'Summary', 'metric': 'Total Services Completed', 'val': serv_count},
            {'category': 'Revenue', 'metric': 'Total Billing Revenue (₹)', 'val': f"₹{rev:,.2f}"}
        ]
        return api_success({'type': 'master', 'rows': rows, 'total': len(rows)})


# ═══════════════════════════════════════════════════════════════
# EXCEL REPORTS EXPORT (.XLS)
# ═══════════════════════════════════════════════════════════════
@app.route('/api/reports/export', methods=['GET'])
def export_report():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    report_type = request.args.get('report_type', 'services') # services, installations, bills, master
    filter_type = request.args.get('filter_type', 'custom')   # custom, month, week, single_date, all
    
    start_date, end_date = parse_report_dates(filter_type, request.args)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{report_type.capitalize()} Report"

    
    # Styling
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    title_font  = Font(name="Arial", size=14, bold=True, color="1F2937")

    ws.append([f"Aqua Care Water Solutions – {report_type.upper()} REPORT"])
    ws.cell(row=1, column=1).font = title_font
    ws.append([f"Filter: {filter_type.upper()} ({start_date.strftime('%d-%b-%Y')} to {end_date.strftime('%d-%b-%Y')})"])
    ws.append([]) # empty row

    if report_type == 'services':
        headers = ["ID", "Customer Name", "Mobile", "Service Type", "Technician", "Service Date", "TDS Before", "TDS After", "Service Charge (₹)", "Parts Total (₹)", "Total Bill (₹)"]
        ws.append(headers)
        services = Service.query.filter(Service.service_date >= start_date, Service.service_date <= end_date).order_by(Service.service_date.desc()).all()
        for s in services:
            c = s.customer
            t = s.technician
            ws.append([
                s.id,
                c.customer_name if c else 'N/A',
                c.mobile if c else 'N/A',
                s.service_type,
                t.technician_name if t else 'N/A',
                s.service_date.strftime('%Y-%m-%d'),
                float(s.tds_before or 0),
                float(s.tds_after or 0),
                float(s.service_charge or 0),
                float(s.parts_total_cost or 0),
                float(s.total_bill or 0)
            ])
    elif report_type == 'installations':
        headers = ["ID", "Customer Name", "Mobile", "Product Name", "Technician", "Installation Date", "Source Water", "Input TDS", "Output TDS", "Cost Price (₹)", "Selling Price (₹)", "Profit (₹)"]
        ws.append(headers)
        insts = Installation.query.filter(Installation.installation_date >= start_date, Installation.installation_date <= end_date).order_by(Installation.installation_date.desc()).all()
        for i in insts:
            c = i.customer
            p = i.product
            t = i.technician
            ws.append([
                i.id,
                c.customer_name if c else 'N/A',
                c.mobile if c else 'N/A',
                p.product_name if p else 'N/A',
                t.technician_name if t else 'N/A',
                i.installation_date.strftime('%Y-%m-%d'),
                i.source_water_type or 'N/A',
                float(i.input_tds or 0),
                float(i.output_tds or 0),
                float(i.cost_price or 0),
                float(i.selling_price or 0),
                float(i.profit or 0)
            ])
    elif report_type == 'bills':
        headers = ["Invoice #", "Customer Name", "Mobile", "Bill Date", "Subtotal (₹)", "Service Charge (₹)", "Grand Total (₹)", "Payment Mode", "Payment Status"]
        ws.append(headers)
        bills = Bill.query.filter(Bill.bill_date >= start_date, Bill.bill_date <= end_date).order_by(Bill.bill_date.desc()).all()
        for b in bills:
            c = b.customer
            ws.append([
                b.invoice_number,
                c.customer_name if c else 'N/A',
                c.mobile if c else 'N/A',
                b.bill_date.strftime('%Y-%m-%d'),
                float(b.subtotal or 0),
                float(b.service_charge or 0),
                float(b.grand_total or 0),
                b.payment_mode,
                b.payment_status
            ])
    else: # master
        headers = ["Category", "Metric / Detail", "Count / Value"]
        ws.append(headers)
        ws.append(["Summary", "Total Services Recorded", Service.query.filter(Service.service_date >= start_date, Service.service_date <= end_date).count()])
        ws.append(["Summary", "Total Installations", Installation.query.filter(Installation.installation_date >= start_date, Installation.installation_date <= end_date).count()])
        ws.append(["Summary", "Total Revenue", sum(float(b.grand_total) for b in Bill.query.filter(Bill.bill_date >= start_date, Bill.bill_date <= end_date).all())])

    # Header styling (row 4)
    header_row_idx = 4
    for cell in ws[header_row_idx]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Auto column width
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"AquaCare_{report_type}_{filter_type}_{start_date.strftime('%Y%m%d')}_to_{end_date.strftime('%Y%m%d')}.xls"
    return send_file(
        output,
        mimetype="application/vnd.ms-excel",
        as_attachment=True,
        download_name=filename
    )



# ═══════════════════════════════════════════════════════════════
# FILE UPLOAD – TECHNICIAN PHOTO
# ═══════════════════════════════════════════════════════════════
import os, uuid
from werkzeug.utils import secure_filename

UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), '..', 'frontend', 'static', 'uploads')
ALLOWED_EXTS  = {'png', 'jpg', 'jpeg', 'gif', 'webp'}

def _allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTS

@app.route('/api/upload/technician-photo', methods=['POST'])
def upload_technician_photo():
    if 'photo' not in request.files:
        return api_error('No file part', 400)
    file = request.files['photo']
    if file.filename == '':
        return api_error('No file selected', 400)
    if not _allowed_file(file.filename):
        return api_error('File type not allowed. Use PNG, JPG, JPEG, GIF or WEBP.', 400)
    # Max 5 MB
    file.seek(0, 2)
    size = file.tell()
    file.seek(0)
    if size > 5 * 1024 * 1024:
        return api_error('File too large. Max 5 MB.', 400)

    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
    ext      = secure_filename(file.filename).rsplit('.', 1)[1].lower()
    filename = f"tech_{uuid.uuid4().hex[:12]}.{ext}"
    filepath = os.path.join(UPLOAD_FOLDER, filename)
    file.save(filepath)
    url = f"/static/uploads/{filename}"
    return api_success({'url': url}, 'Photo uploaded successfully')


# ═══════════════════════════════════════════════════════════════
# THERMAL RECEIPT DATA ENDPOINTS
# ═══════════════════════════════════════════════════════════════

def _build_receipt_data(invoice_no, customer, bill_date_str, line_items, bill_type='ESTIMATE'):
    """Build standardised receipt payload consumed by the frontend print engine."""
    # Use registered user business info if available, else fall back to defaults
    business_info = {
        'name':    'Aqua Care Water Solution',
        'address': 'Erode - 638001',
        'mobile':  '99428 33334',
    }
    try:
        user = User.query.first()
        if user:
            business_info['name']    = user.business_name or business_info['name']
            business_info['address'] = user.business_address or business_info['address']
            business_info['mobile']  = user.mobile or business_info['mobile']
    except Exception:
        pass

    # Compute totals
    total_qty   = sum(item.get('qty') or 0 for item in line_items)
    grand_total = sum((item.get('qty') or 0) * (item.get('rate') or 0) for item in line_items)

    return {
        'bill_type':     bill_type,
        'invoice_no':    invoice_no,
        'bill_date':     bill_date_str,
        'business':      business_info,
        'customer_name': customer.customer_name if customer else 'N/A',
        'customer_mobile': customer.mobile if customer else '',
        'line_items':    line_items,
        'total_qty':     total_qty,
        'grand_total':   grand_total,
        'footer':        'Goods once Sold cannot be Taken Back',
    }


@app.route('/api/bills/<int:bid>/receipt', methods=['GET'])
def bill_receipt(bid):
    b = Bill.query.get_or_404(bid)
    customer = b.customer
    line_items = []

    # If linked to a service, use parts + service charge
    if b.service_id:
        svc = Service.query.get(b.service_id)
        if svc:
            if float(svc.service_charge or 0) > 0:
                line_items.append({'description': 'Service Charge', 'qty': 1, 'rate': float(svc.service_charge)})
            for p in svc.parts:
                line_items.append({
                    'description': p.part_name,
                    'qty':  p.quantity,
                    'rate': float(p.selling_price),
                })

    # If linked to an installation, use product as single line item
    if b.installation_id and not line_items:
        inst = Installation.query.get(b.installation_id)
        if inst and inst.product:
            line_items.append({
                'description': inst.product.product_name,
                'qty':  None,
                'rate': float(inst.selling_price),
            })

    # Fall back to bill totals if no items resolved
    if not line_items:
        if float(b.subtotal or 0) > 0:
            line_items.append({'description': 'Items', 'qty': 1, 'rate': float(b.subtotal)})
        if float(b.service_charge or 0) > 0:
            line_items.append({'description': 'Service Charge', 'qty': 1, 'rate': float(b.service_charge)})

    receipt = _build_receipt_data(
        invoice_no    = b.invoice_number,
        customer      = customer,
        bill_date_str = b.bill_date.strftime('%d/%m/%Y') if b.bill_date else date.today().strftime('%d/%m/%Y'),
        line_items    = line_items,
    )
    return api_success(receipt)


@app.route('/api/installations/<int:iid>/receipt', methods=['GET'])
def installation_receipt(iid):
    inst = Installation.query.get_or_404(iid)
    today = date.today()
    line_items = [{
        'description': inst.product.product_name if inst.product else 'RO Unit',
        'qty':  None,
        'rate': float(inst.selling_price),
    }]
    invoice_no = f"INST-{iid}"
    receipt = _build_receipt_data(
        invoice_no    = invoice_no,
        customer      = inst.customer,
        bill_date_str = inst.installation_date.strftime('%d/%m/%Y') if inst.installation_date else today.strftime('%d/%m/%Y'),
        line_items    = line_items,
    )
    return api_success(receipt)


@app.route('/api/services/<int:sid>/receipt', methods=['GET'])
def service_receipt(sid):
    svc = Service.query.get_or_404(sid)
    today = date.today()
    line_items = []
    if float(svc.service_charge or 0) > 0:
        line_items.append({'description': 'Service Charge', 'qty': 1, 'rate': float(svc.service_charge)})
    for p in svc.parts:
        line_items.append({
            'description': p.part_name,
            'qty':  p.quantity,
            'rate': float(p.selling_price),
        })
    if not line_items:
        line_items.append({'description': svc.service_type + ' Service', 'qty': 1, 'rate': 0})

    invoice_no = f"SVC-{sid}"
    receipt = _build_receipt_data(
        invoice_no    = invoice_no,
        customer      = svc.customer,
        bill_date_str = svc.service_date.strftime('%d/%m/%Y') if svc.service_date else today.strftime('%d/%m/%Y'),
        line_items    = line_items,
    )
    return api_success(receipt)


# ═══════════════════════════════════════════════════════════════
# APP ENTRY POINT
# ═══════════════════════════════════════════════════════════════

def create_app():
    with app.app_context():
        try:
            db.create_all()
            # Auto-migrate columns if upgrading existing database
            for col_query in [
                "ALTER TABLE technicians ADD COLUMN IF NOT EXISTS photo_url VARCHAR(500)",
                "ALTER TABLE products ADD COLUMN IF NOT EXISTS cost_price NUMERIC DEFAULT 0.0",
                "ALTER TABLE products ADD COLUMN IF NOT EXISTS selling_price NUMERIC DEFAULT 0.0"
            ]:
                try:
                    with db.engine.begin() as conn:
                        conn.execute(db.text(col_query))
                except Exception:
                    try:
                        clean_query = col_query.replace(" IF NOT EXISTS", "")
                        with db.engine.begin() as conn:
                            conn.execute(db.text(clean_query))
                    except Exception:
                        pass

            # If database is fresh / empty, auto-seed with default admin login and sample data
            try:
                if User.query.count() == 0:
                    from backend.seed_data import seed
                    seed()
                    print("Initial database auto-seeded successfully.")
            except Exception as seed_err:
                print(f"Notice: Auto-seed check: {seed_err}")
        except Exception as e:
            print(f"Warning: db.create_all() failed: {e}")
    return app



if __name__ == '__main__':
    create_app()
    app.run(debug=True, port=5000, host='0.0.0.0')
